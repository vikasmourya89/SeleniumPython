import openpyxl

wrkBook = openpyxl.load_workbook("PythonDemo.xlsx")
book = wrkBook.active
cell = book.cell(row=1, column=2)
print(cell.value)
book.cell(row=2, column=2).value = "Vikas"
print(book.cell(row=2, column=2).value)
print(book.max_row)

dict = {}

for i in range(1, book.max_row + 1):
    if book.cell(row= i, column=1).value == "TC2":
        for j in range(2, book.max_column + 1):
            dict[book.cell(row=1, column=j).value] = book.cell(row=i, column=j).value

print(dict)
