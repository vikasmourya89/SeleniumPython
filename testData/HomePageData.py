import openpyxl


class HomePageData:
    test_HomePage_Data = [{"name": "vikas", "gender": "Male"}, {"name": "shirin", "gender": "Female"},
                            {"name": "mourya", "gender": "Male"}]

    @staticmethod
    def getTestData(tcName):
        wrkBook = openpyxl.load_workbook("PythonDemo.xlsx")
        book = wrkBook.active
        dict = {}
        for i in range(1, book.max_row + 1):
            if book.cell(row=i, column=1).value == tcName:
                for j in range(2, book.max_column + 1):
                    dict[book.cell(row=1, column=j).value] = book.cell(row=i, column=j).value

        return dict